#!/usr/bin/env python3
"""
Чистит данные из папки «Почва ВКО» (на самом деле — замеры загрязнения атмосферы
по городам с разбивкой по годам/кварталам/месяцам).

Входы:
  - 2014/{1,2,3,4} квартал 14.xlsx       — квартальные
  - 2015/{1..12}.15.xlsx                 — месячные
  - 2016.xlsx, 2018.xlsx ... 2024.xlsx   — годовые
  - 2017 1 полугодие.xlsx                — полугодие

Каждый файл: листы — города. На каждом листе таблица «Примесь × характеристики»,
из которой берём: средняя концентрация (мг/м³) и максимально-разовая.

Выход: public/Качество_воздуха_очищенный.xlsx с листами:
  - "Среднегодовая - <вещество>" (год × город)
  - "Полные данные" (long-format со всеми периодами)
"""
import os
import re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = '/sessions/inspiring-peaceful-bell/mnt/GraphFromData/public/Почва ВКО'
DST  = '/sessions/inspiring-peaceful-bell/mnt/GraphFromData/public/Качество_воздуха_очищенный.xlsx'


# --- Нормализация ---

CITY_ALIASES = {
    'г.усть-каменогорск': 'Өскемен',
    'г. усть-каменогорск': 'Өскемен',
    'г.риддер': 'Риддер',
    'г.семей': 'Семей',
    'п.глубокое': 'Глубокое',
    'г.алтай': 'Алтай',
    'г.шемонаиха': 'Шемонаиха',
}


def norm_city(name):
    s = ' '.join(str(name or '').split()).strip()
    return CITY_ALIASES.get(s.lower(), s)


# Нормализация названий загрязнителей.
# Ключ - регулярка (lowercase), значение - короткое имя.
POLLUTANT_MAP = [
    (r'взвешенные\s*частицы\s*рм\s*-?\s*2[,.]\s*5', 'PM2.5'),
    (r'взвешенные\s*частицы\s*рм\s*-?\s*10', 'PM10'),
    (r'взвешенные\s*частицы\s*\(пыль\)', 'Взвешенные вещества (пыль)'),
    (r'взвешенные\s*частицы(?!\s*рм)', 'Взвешенные вещества (пыль)'),
    (r'взвешенные\s*вещества', 'Взвешенные вещества (пыль)'),
    (r'^\s*взвешенные\s*$', 'Взвешенные вещества (пыль)'),
    (r'диоксид\s*серы', 'SO2 (диоксид серы)'),
    (r'оксид\s*углерода', 'CO (оксид углерода)'),
    (r'диоксид\s*азота', 'NO2 (диоксид азота)'),
    (r'оксид\s*азота', 'NO (оксид азота)'),
    (r'озон', 'O3 (озон)'),
    (r'сероводород', 'H2S (сероводород)'),
    (r'фенол', 'Фенол'),
    (r'формальдегид', 'Формальдегид'),
    (r'серная\s*кислота', 'H2SO4 (серная кислота)'),
    (r'фтористый\s*водород', 'HF (фтористый водород)'),
    (r'хлористый\s*водород', 'HCl (хлористый водород)'),
    (r'\bхлор\b', 'Cl2 (хлор)'),
    (r'бенз\(а\)пирен', 'Бенз(а)пирен'),
    (r'свинец', 'Pb (свинец)'),
    (r'кадмий', 'Cd (кадмий)'),
    (r'медь', 'Cu (медь)'),
    (r'цинк', 'Zn (цинк)'),
    (r'мышьяк', 'As (мышьяк)'),
]


def norm_pollutant(name):
    if not name:
        return None
    s = ' '.join(str(name).split()).strip().lower()
    for pattern, canonical in POLLUTANT_MAP:
        if re.search(pattern, s):
            return canonical
    return None  # неопознанное — пропускаем


# --- Парсинг периода из имени файла ---

def parse_period(filepath):
    """Возвращает (year, period_label, period_type, sort_key)."""
    rel = os.path.relpath(filepath, BASE)
    name = os.path.basename(filepath).lower().replace('.xlsx', '')
    parent = os.path.basename(os.path.dirname(filepath))

    # 2014 квартальные: "1 квартал 14"
    m = re.match(r'^(\d)\s*квартал\s*(\d{2,4})', name)
    if m:
        q = int(m.group(1))
        yr_short = m.group(2)
        year = 2000 + int(yr_short) if len(yr_short) == 2 else int(yr_short)
        return year, f'{year} Q{q}', 'quarter', year * 100 + q

    # 2015 месячные: "1.15", "10.15"
    m = re.match(r'^(\d{1,2})\.(\d{2})$', name)
    if m:
        month = int(m.group(1))
        yr_short = int(m.group(2))
        year = 2000 + yr_short
        return year, f'{year}-{month:02d}', 'month', year * 100 + month

    # "2017 1 полугодие"
    m = re.match(r'^(\d{4})\s*1\s*полугодие', name)
    if m:
        year = int(m.group(1))
        return year, f'{year} H1', 'half', year * 100 + 50

    # Годовой: просто "2016", "2018" и т.п.
    m = re.match(r'^(\d{4})$', name)
    if m:
        year = int(m.group(1))
        return year, f'{year}', 'year', year * 100 + 99

    # Если parent - год (2014, 2015), берём оттуда
    m = re.match(r'^(\d{4})$', parent)
    if m:
        return int(m.group(1)), parent, 'year', int(parent) * 100 + 99

    return None, name, 'unknown', 0


# --- Парсинг одного листа ---

def find_data_section(ws):
    """Находит строку с заголовком 'Примесь' и возвращает индекс следующей за подзаголовком."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for cell in row:
            if cell and isinstance(cell, str) and 'Примесь' in cell:
                # Данные обычно начинаются через 2-3 строки (подзаголовки)
                return i + 2  # после подзаголовка
    return None


STOP_MARKERS = ('иза', 'индекс загрязнения', 'состояние качества', 'наименование водного')


def parse_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '.').replace(' ', '').replace(' ', '')
    if not s or s in ('-', '—', 'н/о', 'нет'):
        return None
    try:
        return float(s)
    except ValueError:
        m = re.match(r'^(-?\d*\.?\d+)', s)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                return None
        return None


def parse_sheet(ws):
    """Возвращает список dict: pollutant, avg, max."""
    start = find_data_section(ws)
    if start is None:
        return []
    records = []
    rows = list(ws.iter_rows(values_only=True))
    for i in range(start, min(start + 30, len(rows))):
        row = rows[i]
        # row[1] — название примеси, row[2] — средняя, row[4] — макс.разовая
        name_raw = row[1] if len(row) > 1 else None
        if name_raw and isinstance(name_raw, str):
            low = name_raw.lower()
            if any(stop in low for stop in STOP_MARKERS):
                break
        poll = norm_pollutant(name_raw)
        if not poll:
            continue
        avg = parse_number(row[2] if len(row) > 2 else None)
        mx  = parse_number(row[4] if len(row) > 4 else None)
        if avg is None and mx is None:
            continue
        records.append({'pollutant': poll, 'avg': avg, 'max': mx})
    return records


# --- Стилизация ---

def style_header(ws):
    f = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', start_color='2563EB')
    for cell in ws[1]:
        cell.font = f
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32


def autosize(ws):
    for col in ws.columns:
        m = 0
        letter = col[0].column_letter
        for c in col:
            v = str(c.value) if c.value is not None else ''
            m = max(m, min(len(v), 36))
        ws.column_dimensions[letter].width = max(10, m + 2)


# --- Главный процесс ---

def collect_all_files(base):
    out = []
    for root, _, files in os.walk(base):
        for fn in files:
            if fn.endswith('.xlsx') and not fn.startswith('~'):
                out.append(os.path.join(root, fn))
    return out


def main():
    files = collect_all_files(BASE)
    print(f"Найдено файлов: {len(files)}")

    # long-format: (year, period, period_type, city, pollutant, avg, max)
    all_records = []
    for fp in sorted(files):
        year, period_label, period_type, sort_key = parse_period(fp)
        if year is None:
            print(f"  SKIP (не понимаю период): {os.path.relpath(fp, BASE)}")
            continue
        try:
            wb = load_workbook(fp, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ERROR {fp}: {e}")
            continue
        for sn in wb.sheetnames:
            city = norm_city(sn)
            if not city:
                continue
            ws = wb[sn]
            recs = parse_sheet(ws)
            for r in recs:
                all_records.append({
                    'year':        year,
                    'period':      period_label,
                    'period_type': period_type,
                    'sort_key':    sort_key,
                    'city':        city,
                    'pollutant':   r['pollutant'],
                    'avg':         r['avg'],
                    'max':         r['max'],
                })

    print(f"Извлечено записей: {len(all_records)}")
    if not all_records:
        return

    # === Агрегация по годам (средняя за год по всем периодам) ===
    yearly = defaultdict(list)  # (year, city, pollutant) → [avg values]
    for r in all_records:
        if r['avg'] is not None:
            yearly[(r['year'], r['city'], r['pollutant'])].append(r['avg'])
    yearly_avg = {k: round(sum(v) / len(v), 4) for k, v in yearly.items()}

    cities  = sorted({r['city'] for r in all_records})
    polls   = sorted({r['pollutant'] for r in all_records})
    years   = sorted({r['year'] for r in all_records})

    out = Workbook()
    out.remove(out.active)

    # --- Лист на каждое вещество: год × город (среднегодовая концентрация) ---
    for poll in polls:
        # Только если есть хотя бы одна запись с avg для этого вещества
        if not any((y, c, poll) in yearly_avg for y in years for c in cities):
            continue
        ws = out.create_sheet(f'Средн. {poll}'[:31])
        ws.append(['Год'] + cities)
        for y in years:
            row = [str(y)]
            for c in cities:
                row.append(yearly_avg.get((y, c, poll)))
            ws.append(row)
        style_header(ws)
        autosize(ws)

    # --- Сводный лист: вещество × город (среднее за последний доступный год) ---
    last_year = max(years)
    ws = out.create_sheet(f'Сводка {last_year}')
    ws.append(['Вещество'] + cities)
    for poll in polls:
        if not any((last_year, c, poll) in yearly_avg for c in cities):
            continue
        row = [poll]
        for c in cities:
            row.append(yearly_avg.get((last_year, c, poll)))
        ws.append(row)
    style_header(ws)
    autosize(ws)

    # --- Полные данные ---
    ws = out.create_sheet('Полные данные')
    ws.append(['Год', 'Период', 'Тип периода', 'Город', 'Вещество',
               'Средняя конц., мг/м³', 'Макс.-разовая, мг/м³'])
    for r in sorted(all_records, key=lambda x: (x['sort_key'], x['city'], x['pollutant'])):
        ws.append([str(r['year']), r['period'], r['period_type'], r['city'],
                   r['pollutant'], r['avg'], r['max']])
    style_header(ws)
    autosize(ws)
    ws.freeze_panes = 'A2'

    out.save(DST)
    print(f"[OK] Сохранено: {DST}")
    print(f"     Годов: {len(years)}, городов: {len(cities)}, веществ: {len(polls)}")
    print(f"     Листов: {len(out.sheetnames)}")
    for n in out.sheetnames:
        ws_n = out[n]
        print(f"  - {n}: {ws_n.max_row-1} строк × {ws_n.max_column} кол")


if __name__ == '__main__':
    main()
