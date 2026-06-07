#!/usr/bin/env python3
"""
Чистит «Радиация.xlsx» → public/Радиация_очищенный.xlsx.

Исходник: 1 лист, очень широкая таблица:
  - годы 2013-2024 (по 7 колонок каждый): проб(вода), α, β, проб(почва), Cs-137, проб(воздух), γ-фон
  - 19 районов ВКО

Часть значений — диапазоны вида "0,12-0,21". Для графиков извлекаем среднее значение
(midpoint), оригинал сохраняем в long-format.
"""
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = '/sessions/inspiring-peaceful-bell/mnt/GraphFromData/public'
SRC  = os.path.join(BASE, 'Радиация ВКО', 'Радиация.xlsx')
DST  = os.path.join(BASE, 'Радиация_очищенный.xlsx')


def parse_value(v):
    """Возвращает (numeric_midpoint, original_string).
    Если значение число — оба совпадают. Если диапазон "0,12-0,21" — среднее и оригинал.
    """
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return float(v), str(v)
    s = str(v).strip().replace(',', '.').replace(' ', '')
    if not s:
        return None, None
    # Диапазон: "0.12-0.21"
    m = re.match(r'^(\d*\.?\d+)-(\d*\.?\d+)$', s)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        return (a + b) / 2, str(v).strip()
    # Просто число
    try:
        return float(s), str(v).strip()
    except ValueError:
        return None, str(v).strip()


def style_header(ws):
    f = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', start_color='2563EB')
    for cell in ws[1]:
        cell.font = f
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40


def autosize(ws):
    for col in ws.columns:
        m = 0
        letter = col[0].column_letter
        for c in col:
            v = str(c.value) if c.value is not None else ''
            m = max(m, min(len(v), 40))
        ws.column_dimensions[letter].width = max(12, m + 2)


# Структура колонок одного года: 7 шт
METRICS = [
    ('Вода',   'Проб взято'),
    ('Вода',   'Альфа-активность, Бк/л'),
    ('Вода',   'Бета-активность, Бк/л'),
    ('Почва',  'Проб взято'),
    ('Почва',  'Цезий-137, Бк/кг'),
    ('Воздух', 'Проб взято'),
    ('Воздух', 'Гамма-фон, мкЗв/час'),
]

YEAR_COLS = {  # год → начальный индекс колонки (0-based)
    2013: 3, 2014: 10, 2015: 17, 2016: 24,
    2017: 32, 2018: 39, 2019: 46, 2020: 53,
    2021: 60, 2022: 67, 2023: 74, 2024: 81,
}

DISTRICTS_ROWS = list(range(5, 24))  # row 5..23 в openpyxl (1-based), кроме 24


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb['Лист1']

    rows = list(ws.iter_rows(values_only=True))

    records = []  # long-format: (district, year, env, metric, value_num, value_orig)
    for r_idx in DISTRICTS_ROWS:
        if r_idx >= len(rows):
            continue
        row = rows[r_idx]
        if row[2] is None:
            continue
        district = str(row[2]).strip()
        for year, base_col in YEAR_COLS.items():
            for offset, (env, metric) in enumerate(METRICS):
                col = base_col + offset
                if col >= len(row):
                    continue
                num, orig = parse_value(row[col])
                if num is None and orig is None:
                    continue
                records.append({
                    'district': district,
                    'year':     year,
                    'env':      env,
                    'metric':   metric,
                    'value':    num,
                    'original': orig,
                })

    # ===== Сохраняем =====
    out = Workbook()
    out.remove(out.active)

    districts = []
    seen = set()
    for r in records:
        if r['district'] not in seen:
            seen.add(r['district'])
            districts.append(r['district'])
    years = sorted({r['year'] for r in records})

    # Листы по средам: каждая комбинация (env × metric) даёт лист
    for env in ['Вода', 'Почва', 'Воздух']:
        # Главный показатель (не «проб»)
        main_metrics = [m for e, m in METRICS if e == env and 'роб' not in m]
        for metric in main_metrics:
            sheet_name = f"{env} - {metric.split(',')[0].strip()}"[:31]
            ws_out = out.create_sheet(sheet_name)
            ws_out.append(['Район'] + [str(y) for y in years])
            for d in districts:
                row_out = [d]
                for y in years:
                    val = next((r['value'] for r in records
                                if r['district'] == d and r['year'] == y
                                and r['env'] == env and r['metric'] == metric), None)
                    row_out.append(val)
                ws_out.append(row_out)
            style_header(ws_out)
            autosize(ws_out)

    # Лист "Полные данные" (long format)
    ws_full = out.create_sheet('Полные данные')
    ws_full.append(['Год', 'Район', 'Среда', 'Показатель', 'Значение', 'Оригинал'])
    for r in records:
        ws_full.append([str(r['year']), r['district'], r['env'], r['metric'],
                        r['value'], r['original']])
    style_header(ws_full)
    autosize(ws_full)
    ws_full.freeze_panes = 'A2'

    out.save(DST)
    print(f"[OK] Сохранено: {DST}")
    print(f"     Районов: {len(districts)}, годов: {len(years)}, всего записей: {len(records)}")


if __name__ == '__main__':
    main()
