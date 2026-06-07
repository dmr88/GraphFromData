#!/usr/bin/env python3
"""
Чистит «обобщ данные по питьевой воде.xlsx» → public/Питьевая_вода_очищенный.xlsx.

Структура исходника: лист 'ВКО', год × (сан-хим: проб, не соотв, %; микробио: проб, не соотв, %)
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = '/sessions/inspiring-peaceful-bell/mnt/GraphFromData/public'
SRC  = os.path.join(BASE, 'Питьевая вода ВКО', 'обобщ данные по питьевой воде.xlsx')
DST  = os.path.join(BASE, 'Питьевая_вода_очищенный.xlsx')


def style_header(ws):
    f = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', start_color='2563EB')
    for cell in ws[1]:
        cell.font = f
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36


def autosize(ws):
    for col in ws.columns:
        m = 0
        letter = col[0].column_letter
        for c in col:
            v = str(c.value) if c.value is not None else ''
            m = max(m, min(len(v), 50))
        ws.column_dimensions[letter].width = max(12, m + 2)


def main():
    src_wb = load_workbook(SRC, data_only=True)
    ws = src_wb['ВКО']
    rows = list(ws.iter_rows(values_only=True))

    # Данные начинаются с row 4. Колонки:
    #   1: год, 2: проб (СХ), 3: не соотв (СХ), 4: % (СХ),
    #   5: проб (МБ), 6: не соотв (МБ), 7: % (МБ)
    records = []
    for r in rows[4:]:
        if r[1] is None: continue
        year_raw = r[1]
        year = str(int(year_raw)) if isinstance(year_raw, (int, float)) and float(year_raw).is_integer() \
               else str(year_raw)
        records.append({
            'period': year,
            'sh_proby': r[2],
            'sh_fail':  r[3],
            'sh_pct':   round((r[4] or 0) * 100, 3) if isinstance(r[4], (int, float)) else None,
            'mb_proby': r[5],
            'mb_fail':  r[6],
            'mb_pct':   round((r[6] / r[5] * 100) if r[5] else 0, 3) if isinstance(r[6], (int, float)) and r[5] else None,
        })

    # ===== Сохраняем =====
    out = Workbook()
    out.remove(out.active)

    # Лист 1: Сводная таблица
    ws1 = out.create_sheet('Сводная')
    ws1.append([
        'Год',
        'Проб (сан-хим)', 'Не соотв (сан-хим)', 'Доля не соотв %, сан-хим',
        'Проб (микробио)', 'Не соотв (микробио)', 'Доля не соотв %, микробио',
    ])
    for r in records:
        ws1.append([r['period'], r['sh_proby'], r['sh_fail'], r['sh_pct'],
                    r['mb_proby'], r['mb_fail'], r['mb_pct']])
    style_header(ws1)
    autosize(ws1)

    # Лист 2: Динамика долей (для тренда)
    ws2 = out.create_sheet('Динамика % не соотв')
    ws2.append(['Год', 'Сан-хим %', 'Микробио %'])
    for r in records:
        ws2.append([r['period'], r['sh_pct'], r['mb_pct']])
    style_header(ws2)
    autosize(ws2)

    # Лист 3: Полные данные (long format)
    ws3 = out.create_sheet('Полные данные')
    ws3.append(['Год', 'Категория', 'Показатель', 'Значение'])
    for r in records:
        for cat, prefix in [('Сан-химические', 'sh'), ('Микробиологические', 'mb')]:
            ws3.append([r['period'], cat, 'Проб исследовано', r[f'{prefix}_proby']])
            ws3.append([r['period'], cat, 'Не соответствует', r[f'{prefix}_fail']])
            ws3.append([r['period'], cat, 'Доля не соотв., %', r[f'{prefix}_pct']])
    style_header(ws3)
    autosize(ws3)
    ws3.freeze_panes = 'A2'

    out.save(DST)
    print(f"[OK] Сохранено: {DST}")
    print(f"     Записей: {len(records)} периодов, {len(records) * 6} long-format строк")


if __name__ == '__main__':
    main()
