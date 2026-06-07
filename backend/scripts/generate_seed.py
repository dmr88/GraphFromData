#!/usr/bin/env python3
"""
Единый seed-генератор. Читает все очищенные xlsx из public/ и формирует
один общий seed.sql, который наполняет БД 5 темами:

  1) Заболеваемость       — заболеваемость_очищенный.xlsx (лист "Полные данные")
  2) Питьевая вода ВКО    — Питьевая_вода_очищенный.xlsx (лист "Полные данные")
  3) Радиация ВКО         — Радиация_очищенный.xlsx (лист "Полные данные")
  4) Атмосфера ВКО        — Атмосфера_ВКО_очищенный.xlsx (лист "Полные данные")
  5) Качество воздуха ВКО — Качество_воздуха_очищенный.xlsx (лист "Полные данные")

Запуск:
    python3 backend/scripts/generate_seed.py > backend/sql/seed.sql
"""
import os
import re
import sys
from openpyxl import load_workbook

PUBLIC = '/sessions/inspiring-peaceful-bell/mnt/GraphFromData/public'

# --- Описание тем ---

TOPICS = [
    {
        'slug': 'morbidity',
        'name': 'Заболеваемость населения',
        'description': 'Случаев заболеваний на 100 000 населения, ВКО, 2013-2024',
        'unit': 'случаев на 100 000',
        'file': 'заболеваемость_очищенный.xlsx',
        'sheet': 'Полные данные',
        'has_city': True,
        'has_age': True,
    },
    {
        'slug': 'water',
        'name': 'Питьевая вода',
        'description': 'Качество питьевой воды по санитарно-химическим и микробиологическим показателям',
        'unit': '%',
        'file': 'Питьевая_вода_очищенный.xlsx',
        'sheet': 'Полные данные',
        'has_city': False,
        'has_age': False,
    },
    {
        'slug': 'radiation',
        'name': 'Радиация',
        'description': 'Радиационное состояние по водным, почвенным и воздушным пробам',
        'unit': None,
        'file': 'Радиация_очищенный.xlsx',
        'sheet': 'Полные данные',
        'has_city': True,
        'has_age': False,
    },
    {
        'slug': 'atmosphere',
        'name': 'Атмосфера (выбросы)',
        'description': 'Выбросы загрязняющих веществ, ИЗА5 и объёмы по городам',
        'unit': 'тыс. т',
        'file': 'Атмосфера_ВКО_очищенный.xlsx',
        'sheet': 'Полные данные',
        'has_city': True,
        'has_age': False,
    },
    {
        'slug': 'air_quality',
        'name': 'Качество воздуха (замеры)',
        'description': 'Среднегодовые концентрации загрязнителей по городам, мг/м³',
        'unit': 'мг/м³',
        'file': 'Качество_воздуха_очищенный.xlsx',
        'sheet': 'Полные данные',
        'has_city': True,
        'has_age': False,
    },
]

AGE_ORDER = {'Всего': 1, 'Взрослые': 2, 'Подростки': 3, 'Дети': 4}


# --- Утилиты ---

def sql_escape(v):
    if v is None:
        return 'NULL'
    s = str(v).replace('\\', '\\\\').replace("'", "''")
    return f"'{s}'"


def parse_period(label):
    """('2013'|'6мес.2024'|'2014 Q2'|'2015-03') -> (year, full, sort_order)"""
    s = str(label).strip()
    if re.match(r'^\d{4}$', s):
        y = int(s)
        return y, 1, y * 100 + 99
    m = re.match(r'^(\d{4})\s*Q(\d)$', s)
    if m:
        return int(m.group(1)), 0, int(m.group(1)) * 100 + int(m.group(2)) * 10
    m = re.match(r'^(\d{4})-(\d{2})$', s)
    if m:
        return int(m.group(1)), 0, int(m.group(1)) * 100 + int(m.group(2))
    m = re.match(r'^(\d{4})\s*H1$', s)
    if m:
        return int(m.group(1)), 0, int(m.group(1)) * 100 + 50
    m = re.search(r'(\d{4})', s)
    if m:
        return int(m.group(1)), 0, int(m.group(1)) * 100 + 5
    return 0, 0, 0


# --- Парсинг каждой темы ---

def read_morbidity():
    """('Год','Категория','Город','Возрастная группа','Значение')"""
    wb = load_workbook(os.path.join(PUBLIC, 'заболеваемость_очищенный.xlsx'),
                       data_only=True, read_only=True)
    ws = wb['Полные данные']
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    out = []
    for r in rows:
        if not r[0]:
            continue
        out.append({
            'period':   str(r[0]).strip(),
            'category': str(r[1]).strip() if r[1] else None,
            'city':     str(r[2]).strip() if r[2] else None,
            'age':      str(r[3]).strip() if r[3] else None,
            'value':    float(r[4]) if isinstance(r[4], (int, float)) else None,
        })
    return [r for r in out if r['category'] and r['value'] is not None]


def read_water():
    """('Год','Категория','Показатель','Значение')"""
    wb = load_workbook(os.path.join(PUBLIC, 'Питьевая_вода_очищенный.xlsx'),
                       data_only=True, read_only=True)
    ws = wb['Полные данные']
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = []
    for r in rows:
        if not r[0]:
            continue
        # Категория = "Сан-химические — Проб исследовано" и т.п.
        cat = f"{r[1]} — {r[2]}" if r[1] and r[2] else None
        out.append({
            'period':   str(r[0]).strip(),
            'category': cat,
            'city':     None,
            'age':      None,
            'value':    float(r[3]) if isinstance(r[3], (int, float)) else None,
        })
    return [r for r in out if r['category'] and r['value'] is not None]


def read_radiation():
    """('Год','Район','Среда','Показатель','Значение','Оригинал')"""
    wb = load_workbook(os.path.join(PUBLIC, 'Радиация_очищенный.xlsx'),
                       data_only=True, read_only=True)
    ws = wb['Полные данные']
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = []
    for r in rows:
        if not r[0]:
            continue
        cat = f"{r[2]} — {r[3]}" if r[2] and r[3] else None  # "Вода — Альфа-активность"
        out.append({
            'period':   str(r[0]).strip(),
            'category': cat,
            'city':     str(r[1]).strip() if r[1] else None,
            'age':      None,
            'value':    float(r[4]) if isinstance(r[4], (int, float)) else None,
            'note':     str(r[5]).strip() if r[5] else None,
        })
    return [r for r in out if r['category'] and r['value'] is not None]


def read_atmosphere():
    """('Год','Показатель','Город','Значение')"""
    wb = load_workbook(os.path.join(PUBLIC, 'Атмосфера_ВКО_очищенный.xlsx'),
                       data_only=True, read_only=True)
    ws = wb['Полные данные']
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = []
    for r in rows:
        if not r[0]:
            continue
        out.append({
            'period':   str(r[0]).strip(),
            'category': str(r[1]).strip() if r[1] else None,
            'city':     str(r[2]).strip() if r[2] else None,
            'age':      None,
            'value':    float(r[3]) if isinstance(r[3], (int, float)) else None,
        })
    return [r for r in out if r['category'] and r['value'] is not None]


def read_air_quality():
    """('Год','Период','Тип периода','Город','Вещество','Средняя','Макс')"""
    wb = load_workbook(os.path.join(PUBLIC, 'Качество_воздуха_очищенный.xlsx'),
                       data_only=True, read_only=True)
    ws = wb['Полные данные']
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = []
    for r in rows:
        if not r[0]:
            continue
        period = str(r[1]).strip() if r[1] else str(r[0]).strip()
        avg = float(r[5]) if isinstance(r[5], (int, float)) else None
        mx  = float(r[6]) if isinstance(r[6], (int, float)) else None
        if avg is None and mx is None:
            continue
        # Если есть только среднее — записываем как value, макс в value_extra
        out.append({
            'period':      period,
            'category':    str(r[4]).strip() if r[4] else None,
            'city':        str(r[3]).strip() if r[3] else None,
            'age':         None,
            'value':       avg if avg is not None else mx,
            'value_extra': mx if (avg is not None and mx is not None) else None,
        })
    return [r for r in out if r['category'] and r['value'] is not None]


READERS = {
    'morbidity':   read_morbidity,
    'water':       read_water,
    'radiation':   read_radiation,
    'atmosphere':  read_atmosphere,
    'air_quality': read_air_quality,
}


# --- Сборка SQL ---

def main():
    # Загружаем данные всех тем
    by_topic = {}
    for t in TOPICS:
        try:
            by_topic[t['slug']] = READERS[t['slug']]()
        except FileNotFoundError as e:
            sys.stderr.write(f"WARNING: пропускаю {t['slug']}: {e}\n")
            by_topic[t['slug']] = []

    # Собираем уникальные периоды, города, возрасты (общие справочники)
    periods = {}     # label -> (year, full, sort_order)
    cities  = []     # порядок появления
    city_seen = set()
    ages    = set()

    for slug, recs in by_topic.items():
        for r in recs:
            if r['period'] not in periods:
                periods[r['period']] = parse_period(r['period'])
            if r.get('city') and r['city'] not in city_seen:
                city_seen.add(r['city'])
                cities.append(r['city'])
            if r.get('age'):
                ages.add(r['age'])

    out = sys.stdout.write
    out('-- ===================================================================\n')
    out(f'-- Seed-данные для проекта Graph From Data ({sum(len(r) for r in by_topic.values())} фактов)\n')
    out('-- Темы: ' + ', '.join(t['slug'] for t in TOPICS) + '\n')
    out('-- ===================================================================\n\n')

    out('SET NAMES utf8mb4;\nSET FOREIGN_KEY_CHECKS = 0;\n')
    out('TRUNCATE TABLE `measurements`;\n')
    out('TRUNCATE TABLE `categories`;\n')
    out('TRUNCATE TABLE `topics`;\n')
    out('TRUNCATE TABLE `periods`;\n')
    out('TRUNCATE TABLE `cities`;\n')
    out('TRUNCATE TABLE `age_groups`;\n')
    out('SET FOREIGN_KEY_CHECKS = 1;\n\n')

    # ---- topics ----
    out('-- Темы\n')
    out('INSERT INTO `topics` (`slug`, `name`, `description`, `unit`, `sort_order`) VALUES\n')
    rows = [
        f'  ({sql_escape(t["slug"])}, {sql_escape(t["name"])}, '
        f'{sql_escape(t["description"])}, {sql_escape(t["unit"])}, {i * 10})'
        for i, t in enumerate(TOPICS)
    ]
    out(',\n'.join(rows) + ';\n\n')

    # ---- periods ----
    out('-- Периоды\n')
    out('INSERT INTO `periods` (`label`, `year_num`, `is_full_year`, `sort_order`) VALUES\n')
    rows = []
    for label, (yr, full, so) in sorted(periods.items(), key=lambda kv: kv[1][2]):
        rows.append(f'  ({sql_escape(label)}, {yr}, {full}, {so})')
    out(',\n'.join(rows) + ';\n\n')

    # ---- cities ----
    out('-- Города\n')
    out('INSERT INTO `cities` (`name`, `is_region_total`, `sort_order`) VALUES\n')
    rows = []
    for i, name in enumerate(cities):
        is_total = 1 if name == 'ВКО (всего)' else 0
        rows.append(f'  ({sql_escape(name)}, {is_total}, {i * 10})')
    out(',\n'.join(rows) + ';\n\n')

    # ---- age_groups ----
    if ages:
        out('-- Возрастные группы\n')
        out('INSERT INTO `age_groups` (`name`, `sort_order`) VALUES\n')
        sorted_ages = sorted(ages, key=lambda n: AGE_ORDER.get(n, 99))
        rows = [
            f'  ({sql_escape(n)}, {AGE_ORDER.get(n, 99) * 10})'
            for n in sorted_ages
        ]
        out(',\n'.join(rows) + ';\n\n')

    # ---- categories ----
    out('-- Категории по темам\n')
    out('INSERT INTO `categories` (`topic_id`, `name`, `sort_order`) VALUES\n')
    cat_rows = []
    for t in TOPICS:
        recs = by_topic[t['slug']]
        cats_seen = []
        for r in recs:
            if r['category'] not in cats_seen:
                cats_seen.append(r['category'])
        for i, name in enumerate(cats_seen):
            cat_rows.append(
                f'  ((SELECT id FROM topics WHERE slug = {sql_escape(t["slug"])}), '
                f'{sql_escape(name)}, {i * 10})'
            )
    out(',\n'.join(cat_rows) + ';\n\n')

    # ---- measurements ----
    out('-- Факты\n')
    BATCH = 400
    all_facts = []
    for t in TOPICS:
        for r in by_topic[t['slug']]:
            all_facts.append((t['slug'], r))
    out(f'-- Всего записей: {len(all_facts)}\n\n')

    for batch_start in range(0, len(all_facts), BATCH):
        batch = all_facts[batch_start:batch_start + BATCH]
        out('INSERT IGNORE INTO `measurements` '
            '(`topic_id`, `period_id`, `category_id`, `city_id`, `age_group_id`, '
            '`value`, `value_extra`, `note`) VALUES\n')
        value_rows = []
        for slug, r in batch:
            city_sub = ('(SELECT id FROM cities WHERE name = '
                        f'{sql_escape(r["city"])})' if r.get('city') else 'NULL')
            age_sub  = ('(SELECT id FROM age_groups WHERE name = '
                        f'{sql_escape(r["age"])})' if r.get('age') else 'NULL')
            note     = sql_escape(r.get('note')) if r.get('note') else 'NULL'
            extra    = (f'{r["value_extra"]:.4f}' if r.get('value_extra') is not None else 'NULL')
            value_rows.append(
                '  ('
                f'(SELECT id FROM topics WHERE slug = {sql_escape(slug)}), '
                f'(SELECT id FROM periods WHERE label = {sql_escape(r["period"])}), '
                f'(SELECT id FROM categories WHERE name = {sql_escape(r["category"])} '
                f'AND topic_id = (SELECT id FROM topics WHERE slug = {sql_escape(slug)})), '
                f'{city_sub}, {age_sub}, '
                f'{r["value"]:.4f}, {extra}, {note}'
                ')'
            )
        out(',\n'.join(value_rows) + ';\n\n')

    out('-- Готово.\n')

    # Отчёт в stderr
    sys.stderr.write('\n=== Сводка ===\n')
    sys.stderr.write(f'Тем: {len(TOPICS)}, периодов: {len(periods)}, '
                     f'городов: {len(cities)}, возрастов: {len(ages)}\n')
    for t in TOPICS:
        n = len(by_topic[t['slug']])
        sys.stderr.write(f"  {t['slug']:12s}: {n:>6} фактов\n")
    sys.stderr.write(f'Итого: {sum(len(r) for r in by_topic.values())} фактов\n')


if __name__ == '__main__':
    main()
